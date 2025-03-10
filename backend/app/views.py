from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from django.contrib.auth import authenticate
from rest_framework_simplejwt.tokens import RefreshToken
from .utils import upload_to_sharepoint
from office365.sharepoint.client_context import ClientContext
from office365.runtime.auth.user_credential import UserCredential
import pandas as pd
import io
import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime
import json
import math
from django.core.serializers.json import DjangoJSONEncoder

class LoginView(APIView):
    def post(self, request):
        username = request.data.get('username')
        password = request.data.get('password')

        user = authenticate(username=username, password=password)
        if user:
            refresh = RefreshToken.for_user(user)
            return Response({
                'refresh': str(refresh),
                'access': str(refresh.access_token),
            }, status=status.HTTP_200_OK)
        else:
            return Response({'error': 'Credenciales inválidas'}, status=status.HTTP_401_UNAUTHORIZED)

class SKULoadView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            # Configurar SharePoint
            sharepoint_url = "https://agrosuper.sharepoint.com/sites/PanelPlantaRosario"
            folder_path = "/sites/PanelPlantaRosario/Documentos compartidos/1.- Torre de Control/1.- Gestión TC/2- Registro Bitácora TC (interfaz web)"
            archivo_excel = 'SKU.xlsx'

            ctx = ClientContext(sharepoint_url).with_credentials(
                UserCredential('aialvarado@agrosuper.com', 'Produccion2025.')
            )

            # Leer el archivo existente desde SharePoint
            response = ctx.web.get_file_by_server_relative_url(folder_path + "/" + archivo_excel).execute_query()
            file_content = io.BytesIO()
            response.download(file_content).execute_query()
            file_content.seek(0)

            # Leer el archivo Excel
            df_sku = pd.read_excel(file_content)

            # Verificar que el archivo tenga las columnas esperadas
            if 'SKU' not in df_sku.columns or 'Producto' not in df_sku.columns:
                return Response({'error': 'El archivo SKU.xlsx no tiene el formato esperado.'}, status=status.HTTP_400_BAD_REQUEST)

            # Convertir DataFrame a lista de diccionarios
            sku_data = df_sku.to_dict('records')

            return Response({'skuData': sku_data}, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class FormularioView(APIView):
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        try:
            data = request.data
            sharepoint_url = "https://agrosuper.sharepoint.com/sites/PanelPlantaRosario"
            folder_path = "/sites/PanelPlantaRosario/Documentos compartidos/1.- Torre de Control/1.- Gestión TC/2- Registro Bitácora TC (interfaz web)"
            archivo_excel = 'Bitácora TC.xlsx'

            ctx = ClientContext(sharepoint_url).with_credentials(
                UserCredential('aialvarado@agrosuper.com', 'Produccion2025.')
            )

            # Leer el archivo existente desde SharePoint
            response = ctx.web.get_file_by_server_relative_url(folder_path + "/" + archivo_excel).execute_query()
            file_content = io.BytesIO()
            response.download(file_content).execute_query()
            file_content.seek(0)

            # Leer todas las hojas del archivo Excel
            xls = pd.ExcelFile(file_content)
            df_existente = pd.read_excel(xls, sheet_name='Hoja1')  # Asumiendo que los registros están en 'Hoja1'
            df_sku = pd.read_excel(xls, sheet_name='SKU')  # Leer la hoja SKU

            user = request.user
            full_name = f"{user.first_name} {user.last_name}"
            sku = data.get('sku')
            # Crear nuevo registro con fórmula VLOOKUP para el producto
            nuevo_registro = pd.DataFrame([{
                "Fecha y Hora": datetime.now().strftime("%d-%m-%Y %H:%M:%S"),
                "Tipo Notificación": data.get('tipo_notificacion'),
                "Sentido": data.get('sentido'),
                "Canal de Comunicación": data.get('canal_comunicacion'),
                "Emisor": full_name,
                "Área": data.get('area'),
                "Unidad": data.get('unidad'),
                "Indicador": data.get('indicador').title(),
                "Valor %": data.get('valor'),
                "Desviación": data.get('desviacion'),
                "Hora de Desviación": data.get('hora_desviacion'),
                "Respuesta": data.get('respuesta'),
                "SKU": data.get('sku'),
                "Producto": f'=VLOOKUP({sku},SKU!A:B,2,FALSE)',
                "Receptor": data.get('receptor').title(),
                "Observaciones": data.get('observaciones')
            }])

            # Combinar los datos existentes con el nuevo registro
            df_final = pd.concat([df_existente, nuevo_registro], ignore_index=True)

            # Crear un nuevo archivo Excel en memoria
            excel_file = io.BytesIO()
            with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                df_final.to_excel(writer, sheet_name='Hoja1', index=False)
                df_sku.to_excel(writer, sheet_name='SKU', index=False)

            # Cargar el archivo para aplicar formato
            excel_file.seek(0)
            workbook = openpyxl.load_workbook(excel_file)
            worksheet = workbook['Hoja1']  # Trabajar con la hoja principal

            # Aplicar formatos (el código existente de formato)
            desviacion_colors = {
                'D1': 'C6EFCE',
                'D2': 'FDE208',
                'D3': 'FFA500',
                'D4': 'FF0000'
            }

            desv_col = None
            for idx, col in enumerate(df_final.columns):
                if col == 'Desviación':
                    desv_col = idx + 1
                    break

            if desv_col is not None:
                for row in range(2, len(df_final) + 2):
                    cell = worksheet.cell(row=row, column=desv_col)
                    if cell.value in desviacion_colors:
                        cell.fill = PatternFill(
                            start_color=desviacion_colors[cell.value],
                            end_color=desviacion_colors[cell.value],
                            fill_type='solid'
                        )

            # Autoajustar columnas
            for col in worksheet.columns:
                max_length = 0
                column = openpyxl.utils.get_column_letter(col[0].column)
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column].width = adjusted_width

            # Guardar el archivo con los cambios
            excel_file = io.BytesIO()
            workbook.save(excel_file)
            excel_file.seek(0)

            # Subir el archivo actualizado a SharePoint
            target_folder = ctx.web.get_folder_by_server_relative_url(folder_path)
            target_file = target_folder.upload_file(archivo_excel, excel_file.read()).execute_query()

            return Response({'message': 'Registro guardado exitosamente'}, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class UserProfileView(APIView):
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        user = request.user
        data = {
            'first_name': user.first_name,
            'last_name': user.last_name,
            'email': user.email,
        }
        return Response(data)

class BitacoraDataView(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        try:
            # Configurar SharePoint
            sharepoint_url = "https://agrosuper.sharepoint.com/sites/PanelPlantaRosario"
            folder_path = "/sites/PanelPlantaRosario/Documentos compartidos/1.- Torre de Control/1.- Gestión TC/2- Registro Bitácora TC (interfaz web)"
            archivo_excel = 'Bitácora TC.xlsx'
            ctx = ClientContext(sharepoint_url).with_credentials(
                UserCredential('aialvarado@agrosuper.com', 'Produccion2025.')
            )
            # Leer el archivo existente desde SharePoint
            response = ctx.web.get_file_by_server_relative_url(folder_path + "/" + archivo_excel).execute_query()
            file_content = io.BytesIO()
            response.download(file_content).execute_query()
            file_content.seek(0)
            
            # Leer el archivo Excel
            df = pd.read_excel(file_content)
            
            # Validar que existan las columnas requeridas
            required_columns = ['Fecha y Hora', 'Desviación', 'Área', 'Indicador']
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                return Response(
                    {'error': f'Faltan columnas requeridas: {", ".join(missing_columns)}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Formatear y limpiar datos
            df['Fecha y Hora'] = pd.to_datetime(
                df['Fecha y Hora'], 
                format='%d-%m-%Y %H:%M:%S', 
                errors='coerce'
            ).dt.strftime('%d-%m-%Y %H:%M:%S')
            
            # Limpiar y normalizar los datos
            df = df.fillna('')
            df['Indicador'] = df['Indicador'].astype(str).str.strip()
            df['Área'] = df['Área'].astype(str).str.strip()
            df['Desviación'] = df['Desviación'].astype(str).str.strip()

            # Preparar estadísticas adicionales
            indicadores_stats = []
            for indicador in df['Indicador'].unique():
                if not indicador:  # Skip empty indicators
                    continue
                    
                mask = df['Indicador'] == indicador
                areas = df[mask]['Área'].unique()
                total = mask.sum()
                
                stats = {
                    'indicador': indicador,
                    'areas': ', '.join(filter(None, areas)),  # Filter out empty areas
                    'total': int(total),
                    'D1': int(df[mask & (df['Desviación'] == 'D1')].shape[0]),
                    'D2': int(df[mask & (df['Desviación'] == 'D2')].shape[0]),
                    'D3': int(df[mask & (df['Desviación'] == 'D3')].shape[0]),
                    'D4': int(df[mask & (df['Desviación'] == 'D4')].shape[0])
                }
                indicadores_stats.append(stats)
            
            # Ordenar por total y tomar los top 10
            indicadores_stats.sort(key=lambda x: x['total'], reverse=True)
            top_indicadores = indicadores_stats[:10]
            
            # Convertir DataFrame a lista de diccionarios para los datos principales
            data = df.to_dict(orient='records')
            
            return Response({
                'data': data,
                'stats': {
                    'total_registros': len(data),
                    'total_desviaciones': len(df[df['Desviación'].isin(['D1', 'D2', 'D3', 'D4'])]),
                    'conteo_por_desviacion': {
                        desv: int(df['Desviación'].eq(desv).sum())
                        for desv in ['D1', 'D2', 'D3', 'D4']
                    }
                },
                'top_indicadores': top_indicadores
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class CargaMasivaView(APIView):
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        try:
            registros = request.data.get('registros', [])
            if not registros:
                return Response({'error': 'No se proporcionaron registros'}, status=status.HTTP_400_BAD_REQUEST)

            sharepoint_url = "https://agrosuper.sharepoint.com/sites/PanelPlantaRosario"
            folder_path = "/sites/PanelPlantaRosario/Documentos compartidos/1.- Torre de Control/1.- Gestión TC/2- Registro Bitácora TC (interfaz web)"
            archivo_excel = 'Bitácora TC.xlsx'

            ctx = ClientContext(sharepoint_url).with_credentials(
                UserCredential('aialvarado@agrosuper.com', 'Produccion2025.')
            )

            # Leer el archivo existente
            response = ctx.web.get_file_by_server_relative_url(folder_path + "/" + archivo_excel).execute_query()
            file_content = io.BytesIO()
            response.download(file_content).execute_query()
            file_content.seek(0)

            xls = pd.ExcelFile(file_content)
            df_existente = pd.read_excel(xls, sheet_name='Hoja1')
            df_sku = pd.read_excel(xls, sheet_name='SKU')

            user = request.user
            full_name = f"{user.first_name} {user.last_name}"

            nuevos_registros = []
            for registro in registros:
                nuevo_registro = {
                    "Fecha y Hora": datetime.now().strftime("%d-%m-%Y %H:%M:%S"),
                    "Tipo Notificación": registro['tipo_notificacion'],
                    "Sentido": registro['sentido'],
                    "Canal de Comunicación": registro['canal_comunicacion'],
                    "Emisor": full_name,
                    "Área": registro['area'],
                    "Unidad": registro['unidad'],
                    "Indicador": registro['indicador'],
                    "Valor %": registro['valor'],
                    "Desviación": registro['desviacion'],
                    "Hora de Desviación": registro['hora_desviacion'],
                    "Respuesta": registro['respuesta'],
                    "SKU": registro['sku'],
                     "Producto": registro['producto'].title() if registro['producto'] else f'=VLOOKUP({registro["sku"]},SKU!A:B,2,FALSE)',
                    "Receptor": registro['receptor'].title(),
                    "Observaciones": registro['observaciones']
                }
                nuevos_registros.append(nuevo_registro)

            df_nuevos = pd.DataFrame(nuevos_registros)
            df_final = pd.concat([df_existente, df_nuevos], ignore_index=True)

            # Crear nuevo archivo Excel
            excel_file = io.BytesIO()
            with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                df_final.to_excel(writer, sheet_name='Hoja1', index=False)
                df_sku.to_excel(writer, sheet_name='SKU', index=False)

            # Aplicar formatos
            excel_file.seek(0)
            workbook = openpyxl.load_workbook(excel_file)
            worksheet = workbook['Hoja1']

            desviacion_colors = {
                'D1': 'C6EFCE',
                'D2': 'FDE208',
                'D3': 'FFA500',
                'D4': 'FF0000'
            }

            desv_col = None
            for idx, col in enumerate(df_final.columns):
                if col == 'Desviación':
                    desv_col = idx + 1
                    break

            if desv_col is not None:
                for row in range(2, len(df_final) + 2):
                    cell = worksheet.cell(row=row, column=desv_col)
                    if cell.value in desviacion_colors:
                        cell.fill = PatternFill(
                            start_color=desviacion_colors[cell.value],
                            end_color=desviacion_colors[cell.value],
                            fill_type='solid'
                        )

            # Autoajustar columnas
            for col in worksheet.columns:
                max_length = 0
                column = openpyxl.utils.get_column_letter(col[0].column)
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column].width = adjusted_width

            # Guardar y subir archivo
            excel_file = io.BytesIO()
            workbook.save(excel_file)
            excel_file.seek(0)

            target_folder = ctx.web.get_folder_by_server_relative_url(folder_path)
            target_file = target_folder.upload_file(archivo_excel, excel_file.read()).execute_query()

            return Response({'message': f'Se guardaron {len(registros)} registros exitosamente'}, 
                          status=status.HTTP_200_OK)

        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class ExcelDataView(APIView):
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        try:
            # SharePoint configuration
            sharepoint_url = "https://agrosuper.sharepoint.com/sites/PanelPlantaRosario"
            folder_path = "/sites/PanelPlantaRosario/Documentos compartidos/1.- Torre de Control/1.- Gestión TC/2- Registro Bitácora TC (interfaz web)"
            archivo_excel = 'Bitácora TC.xlsx'

            # Initialize SharePoint context
            ctx = ClientContext(sharepoint_url).with_credentials(
                UserCredential('aialvarado@agrosuper.com', 'Produccion2025.')
            )

            # Get file from SharePoint
            response = ctx.web.get_file_by_server_relative_url(folder_path + "/" + archivo_excel).execute_query()
            file_content = io.BytesIO()
            response.download(file_content).execute_query()
            file_content.seek(0)

            # Read Excel file
            df = pd.read_excel(file_content)
            
            # Convert datetime columns to string format
            for col in df.columns:
                if pd.api.types.is_datetime64_any_dtype(df[col]):
                    df[col] = df[col]

            # Handle NaN values
            df = df.fillna('')

            # Convert to list of dictionaries with row IDs
            data = [
                {**row, 'id': idx} 
                for idx, row in enumerate(df.to_dict('records'))
            ]

            return Response({
                'status': 'success',
                'data': data,
                'columns': list(df.columns)
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                'status': 'error',
                'message': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)